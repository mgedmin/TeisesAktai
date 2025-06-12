#!/usr/bin/env -S uv run --script
# /// script
# dependencies = [
#     "openpyxl",
# ]
# ///
"""
XLS failo visuose worksheetuose raskime unikalius teisės aktų identifikacinius
numerius.

XLS turi turėti stulpelį "Pavadinimas", o jame turi būti tekstas
'Identifikacinis kodas YYYY-NNNNN'.
"""

import argparse
import datetime
from dataclasses import dataclass
from typing import Iterator

import openpyxl


@dataclass
class Aktas:
    rūšis: str
    pavadinimas: str
    nuoroda: str
    įstaigos_suteiktas_nr: str | int | datetime.datetime
    priėmimo_data: datetime.datetime
    įsigaliojimo_data: datetime.datetime
    # galiojimas/pakeitimai/projektai neįdomu
    priėmė: str = ''
    užregistruota: str = ''

    @property
    def identifikacinis_kodas(self) -> str:
        return self.užregistruota.partition('Identifikacinis kodas ')[-1]


def parse_worksheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
) -> Iterator[Aktas]:
    columns = next(ws.values)  # 1st row
    assert columns[:6] == (
        'Eil. Nr.',
        'Rūšis',
        'Pavadinimas',
        'Įstaigos suteiktas Nr.',
        'Priėmimo data',
        'Įsigaliojimo data',
    )
    aktas = None
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value:
            if aktas is not None:
                yield aktas
            assert isinstance(row[0].value, (int, float))
            assert isinstance(row[1].value, str)
            assert isinstance(row[2].value, str)
            assert row[2].hyperlink is not None
            assert row[2].hyperlink.target is not None
            assert isinstance(row[3].value, (str, int, datetime.datetime)), (
                f'D{idx}: {row[3].value}'
            )
            assert isinstance(row[4].value, datetime.datetime)
            assert isinstance(row[5].value, datetime.datetime)
            aktas = Aktas(
                rūšis=row[1].value,
                pavadinimas=row[2].value,
                nuoroda=row[2].hyperlink.target,
                įstaigos_suteiktas_nr=row[3].value,
                priėmimo_data=row[4].value,
                įsigaliojimo_data=row[5].value,
            )
        elif row[2].value:
            assert aktas is not None
            s = row[2].value
            assert isinstance(s, str), s
            if s.startswith(('Priėmė', 'Enacted')):
                aktas.priėmė = s
            elif s.startswith(('Užregistruota', 'Identifikacinis kodas')):
                aktas.užregistruota = s
            else:
                assert False, f'C{idx}: {s}'
    if aktas is not None:
        yield aktas


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('filename', nargs='?', default='DI.xlsx')
    args = parser.parse_args()

    seen = {}
    wb = openpyxl.load_workbook(args.filename)
    for ws in wb.worksheets:
        name = ws.title
        codes = {
            aktas.identifikacinis_kodas: aktas
            for aktas in parse_worksheet(ws)
        }
        new = set(codes) - set(seen)
        print(f"{name}: aktų - {len(codes)}, naujų - {len(new)}")
        seen.update(codes)

    print(f"Viso: {len(seen)}")
    print(*sorted(seen), sep=', ')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        'Eil. Nr.',
        'Rūšis',
        'Pavadinimas',
        'Priėmė',
        'Įstaigos suteiktas Nr.',
        'Priėmimo data',
        'Įsigaliojimo data',
        'Užregistruota, Identifikacinis kodas',
    ])
    for n, aktas in enumerate(seen.values(), 1):
        ws.append([
            n,
            aktas.rūšis,
            aktas.pavadinimas,
            aktas.priėmė,
            aktas.įstaigos_suteiktas_nr,
            aktas.priėmimo_data,
            aktas.įsigaliojimo_data,
            aktas.užregistruota,
        ])
        ws.cell(row=n, column=3).hyperlink = aktas.nuoroda
    filename = 'DI aktai be dublikatų.xlsx'
    wb.save(filename)
    print(f'Wrote {filename}')


if __name__ == '__main__':
    main()
